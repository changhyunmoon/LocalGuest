from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_rows():
    return [
        # 프로젝트/플랫폼
        ("1", "프로젝트 기반", "-", "멀티모듈 백엔드 구조", "Gradle 멀티모듈(api-server, domain, chat, ai, ai-integration, common) 구성", "완료", "상", "backend/settings.gradle, backend/build.gradle", "아키텍처"),
        ("1.1", "프로젝트 기반", "-", "프론트 SPA 라우팅 구조", "공통/인증/가이드/마이페이지 라우트 분리", "완료", "상", "frontend/src/App.jsx, frontend/src/routes/*", "아키텍처"),
        ("1.2", "프로젝트 기반", "-", "CI/CD 파이프라인", "frontend/backend blue-green 배포 자동화 + k6 배포 워크플로", "완료", "상", ".github/workflows/*.yml", "운영"),
        ("1.3", "프로젝트 기반", "-", "컨테이너 배포 스크립트", "EC2 배포 스크립트 + compose 기반 health-check/전환", "완료", "상", "backend/deployment/deploy.sh, frontend/deployment/deploy.sh", "운영"),
        # F01 회원/인증
        ("2", "F01 회원/인증", "F01-01", "회원가입", "여행자/가이드 가입, 이메일 인증코드 발송/확인, 닉네임 중복 확인", "완료", "상", "/members/join, /members/email-verification/*", "Auth"),
        ("2.1", "F01 회원/인증", "F01-02", "로그인/로그아웃", "JWT 로그인/로그아웃/토큰 재발급", "완료", "상", "/auth/login, /auth/logout, /auth/reissue", "Auth"),
        ("2.2", "F01 회원/인증", "F01-03", "계정 찾기", "아이디 찾기/비밀번호 재설정 메일 발송/검증", "완료", "중", "/members/find-id, /members/password-reset/*", "Member"),
        ("2.3", "F01 회원/인증", "F01-04", "소셜 로그인", "Google OAuth2 콜백 처리", "완료", "중", "SecurityConfig OAuth2, frontend oauth2/callback", "Auth"),
        ("2.4", "F01 회원/인증", "F01-05", "마이페이지 계정정보", "내 프로필 조회/수정, 회원 탈퇴", "완료", "중", "/members/me/profile, /members/me", "Member"),
        # F02 메인/탐색
        ("3", "F02 메인/탐색", "F02-01", "메인 홈", "히어로 배너 + 목적지 + 로컬 전문가 섹션", "완료", "상", "frontend/src/pages/HomePage.jsx", "Frontend"),
        ("3.1", "F02 메인/탐색", "F02-02", "가이드 목록", "지역 필터 기반 목록/카드 노출", "완료", "상", "frontend/src/pages/GuideListPage.jsx", "Frontend"),
        ("3.2", "F02 메인/탐색", "F02-03", "가이드 상세", "프로필/리뷰/매칭 요청 진입", "완료", "상", "frontend/src/pages/GuideDetailPage.jsx", "Frontend"),
        ("3.3", "F02 메인/탐색", "F02-04", "피드 미리보기 정책", "매칭 전/후 노출 분리 정책 반영", "부분", "중", "GuideFeed API + 상세 페이지 정책", "Guide"),
        # F03 AI 매칭
        ("4", "F03 AI 매칭", "F03-01", "AI 컨셉 입력", "자연어 프롬프트 입력/유효성 처리", "완료", "상", "frontend/src/pages/AiQuickSearchPage.jsx", "AI"),
        ("4.1", "F03 AI 매칭", "F03-02", "AI 추천 엔진", "프롬프트 파싱 + 룰 기반 스코어링 + 상위 추천", "완료", "상", "module-ai PromptParser/MatchingEngine", "AI"),
        ("4.2", "F03 AI 매칭", "F03-02", "2단계 추천 구조", "DB 후보 선필터링 후 AI 계산으로 비용/속도 최적화", "완료", "상", "module-ai-integration DbBackedGuideCandidateProvider", "AI"),
        ("4.3", "F03 AI 매칭", "F03-03", "추천 클릭 신호 수집", "추천 클릭/노출 집계 및 추천 보정 반영", "완료", "중", "/ai/recommend/click, AiRecommendClickStore", "AI"),
        ("4.4", "F03 AI 매칭", "F03-04", "매칭 요청 전송", "추천 결과에서 가이드 매칭 요청 생성", "완료", "상", "/matching/requests", "Matching"),
        # F04 채팅
        ("5", "F04 채팅", "F04-01", "채팅방 오케스트레이션", "매칭 상태 기반 채팅방 생성/조회", "완료", "상", "ChatOrchestrationController", "Chat"),
        ("5.1", "F04 채팅", "F04-02", "실시간 메시징", "SockJS + STOMP 연결, 발행/구독, 읽음 처리", "완료", "상", "WebSocketConfig, MessagesPage", "Chat"),
        ("5.2", "F04 채팅", "F04-03", "알림 스트림", "SSE 구독 기반 알림 채널", "완료", "중", "/notifications/subscribe", "Chat"),
        ("5.3", "F04 채팅", "F04-03", "안심번호/통화 연동", "외부 통신수단 연동", "미구현", "하", "요구사항 명세 기준", "Chat"),
        # F05 결제/예약
        ("6", "F05 예약/결제", "F05-01", "매칭 제안/수락/거절", "요청-제안-수락/거절 상태 전이", "완료", "상", "MatchRequestController", "Matching"),
        ("6.1", "F05 예약/결제", "F05-02", "결제 생성/확정", "결제 생성, KakaoPay/FakePG 승인 처리", "완료", "상", "PaymentController, PaymentService", "Matching"),
        ("6.2", "F05 예약/결제", "F05-03", "일정 연장", "연장 선택/마감 스케줄링", "완료", "중", "TourExtensionController, MatchingSchedulingConfig", "Matching"),
        ("6.3", "F05 예약/결제", "F05-04", "환불 처리", "환불 요청/처리 상태 관리", "완료", "중", "/matching/payments/refunds", "Matching"),
        ("6.4", "F05 예약/결제", "-", "결제 리다이렉트 UX", "KakaoPay redirect 성공/취소/실패 stub 처리", "완료", "중", "KakaoPayRedirectController, PaymentKakaoStubPage", "Matching"),
        # F06 가이드
        ("7", "F06 가이드 기능", "F06-01", "가이드 프로필 관리", "가이드 프로필 조회/수정/승인/활성화", "완료", "상", "GuideProfileController", "Guide"),
        ("7.1", "F06 가이드 기능", "F06-02", "경력/소개 관리", "경력 CRUD + 소개/키워드 반영", "완료", "중", "GuideCareerController", "Guide"),
        ("7.2", "F06 가이드 기능", "F06-03", "피드/이미지 관리", "피드 CRUD, 이미지 업로드/삭제", "완료", "상", "GuideFeedController, GuideImageController", "Guide"),
        ("7.3", "F06 가이드 기능", "F06-04", "스케줄 관리", "가용일/상태/블락/예약 상태 동기화", "완료", "상", "GuideScheduleController", "Guide"),
        ("7.4", "F06 가이드 기능", "F06-05", "정산 예상 금액", "가이드 정산 예상 금액 조회", "완료", "중", "/guides/{guideId}/settlement/expected", "Guide"),
        ("7.5", "F06 가이드 기능", "F06-06", "가이드 설정/관리 화면", "가이드 마이페이지(프로필/피드/리뷰/정산/설정)", "완료", "중", "frontend/src/pages/Guide*Page.jsx", "Frontend"),
        # F07 리뷰/마이페이지
        ("8", "F07 리뷰/마이페이지", "F07-01", "리뷰 작성/조회/삭제", "투어 후 리뷰 작성, 내 리뷰/가이드 리뷰 조회", "완료", "상", "ReviewController", "Review"),
        ("8.1", "F07 리뷰/마이페이지", "F07-02", "별점 반영", "가이드 평점 업데이트 반영", "완료", "중", "/guides/{guideId}/rating", "Review"),
        ("8.2", "F07 리뷰/마이페이지", "-", "여행자 마이페이지", "스크랩북/일정/결제/프라이버시/투어/예정 일정", "완료", "중", "mypageRoutes + GuestUpcomingTripsPage", "Frontend"),
        ("8.3", "F07 리뷰/마이페이지", "-", "지도 기반 여행 기록", "카카오맵 기반 코스/티켓 상세 시각화", "완료", "중", "MypageScrapbook*Page, GuideMatchedCoursePage", "Frontend"),
        # 품질/운영
        ("9", "품질/운영", "-", "API 문서화", "Springdoc OpenAPI + Swagger UI", "완료", "중", "springdoc-openapi, OpenApiSecurityConfig", "Backend"),
        ("9.1", "품질/운영", "-", "로깅/관측", "logback 파일 롤링 + traceId + actuator health", "완료", "상", "logback-spring.xml, /actuator/health", "Observability"),
        ("9.2", "품질/운영", "-", "부하 테스트", "k6 스크립트 배포 + stress-test endpoint", "완료", "중", "stress-test/scripts/test1.js, StressTestController", "Performance"),
        ("9.3", "품질/운영", "-", "테스트 고도화", "도메인 단위 테스트 확장/회귀 자동화 강화", "부분", "중", "module-* test 코드 존재", "QA"),
    ]


def write_excel(output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"

    headers = [
        "WBS",
        "기능그룹",
        "기능ID",
        "작업명",
        "세부설명(코드우선 기준)",
        "현재상태",
        "우선순위",
        "근거(코드/문서)",
        "영역",
    ]
    ws.append(headers)

    rows = build_rows()
    for row in rows:
        ws.append(row)

    # Header style
    fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for c in ws[1]:
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Row style
    status_fill = {
        "완료": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "부분": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "미구현": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status = row[5].value
        if status in status_fill:
            row[5].fill = status_fill[status]
        row[0].alignment = Alignment(horizontal="center", vertical="top")
        row[2].alignment = Alignment(horizontal="center", vertical="top")
        row[5].alignment = Alignment(horizontal="center", vertical="top")
        row[6].alignment = Alignment(horizontal="center", vertical="top")
        row[8].alignment = Alignment(horizontal="center", vertical="top")

    # Column widths
    widths = {
        "A": 8,
        "B": 18,
        "C": 11,
        "D": 28,
        "E": 56,
        "F": 10,
        "G": 10,
        "H": 52,
        "I": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    # 기준 시트
    guide = wb.create_sheet("기준_설명")
    guide.append(["항목", "설명"])
    guide.append(["작성 원칙", "요구사항명세서 기능축(F01~F07)을 기반으로, 현재 코드 구현 여부를 우선 반영"])
    guide.append(["상태 기준", "완료: 코드/화면/엔드포인트 존재, 부분: 일부 흐름 또는 정책 미완성, 미구현: 명세 대비 코드 부재"])
    guide.append(["우선순위 기준", "상: 핵심 유저 플로우, 중: 운영/확장 핵심, 하: 선택 기능"])
    guide.append(["참고 문서", "Team6_2차프로젝트_LocalGuest_요구사항명세서.xlsx"])
    guide.append(["참고 코드", "backend/*, frontend/*, .github/workflows/*, stress-test/*"])
    for c in guide[1]:
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 120
    guide.freeze_panes = "A2"
    for row in guide.iter_rows(min_row=2, max_row=guide.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    out = root / "LocalGuest_WBS_current_code.xlsx"
    write_excel(out)
    print(f"Created: {out}")
